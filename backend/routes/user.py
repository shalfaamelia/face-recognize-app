from flask import Blueprint, request, jsonify, send_from_directory
from werkzeug.utils import secure_filename
import os
import re
import shutil
import subprocess
import uuid
import zipfile
import hashlib
from xml.etree import ElementTree

from db import get_db_connection
from utils.auth_guard import auth_required

user_bp = Blueprint('users', __name__)

# ===============================
# CONFIG
# ===============================
BASE_DIR = os.path.dirname(__file__)
UPLOAD_FOLDER = os.path.join(BASE_DIR, '..', 'dataset')

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}
ALLOWED_ROLES = {'kepala_lab', 'teknisi', 'dosen', 'sarpras', 'mahasiswa'}

USER_IMPORT_REQUIRED_FIELDS = ['nama', 'role']

USER_IMPORT_HEADER_ALIASES = {
    'nama': 'nama',
    'name': 'nama',

    'role': 'role',
    'jabatan': 'role',

    'nim': 'nim',
    'nip': 'nip',

    'prodi': 'prodi',
    'program studi': 'prodi',

    'kelas': 'kelas',

    'email': 'email',
    'password': 'password',

    'foto': 'foto',
    'photo': 'foto',
    'file foto': 'foto',
    'nama foto': 'foto'
}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)


# ===============================
# HELPER
# ===============================
def allowed_file(filename):
    return (
        '.' in filename
        and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
    )

def normalize_import_header(value):
    return str(value or '').strip().lower().replace('\n', ' ')

def column_index(cell_reference):
    letters = ''.join(re.findall(r'[A-Z]+', cell_reference.upper()))
    index = 0

    for char in letters:
        index = index * 26 + (ord(char) - ord('A') + 1)

    return index - 1

def generate_face_label(nama):
    if not nama:
        return "user"

    label = nama.strip().lower()
    label = re.sub(r'[^a-z0-9\s_-]', '', label)
    label = re.sub(r'\s+', '_', label)

    return label or "user"

def get_uploaded_files():
    files = request.files.getlist('files')

    if not files:
        for key in request.files:
            files.extend(request.files.getlist(key))

    return [file for file in files if file and file.filename]

def hash_password_md5(password):
    if not password:
        return None

    password = str(password)

    # Jika sudah terlihat seperti MD5, jangan hash ulang
    if re.fullmatch(r'[a-fA-F0-9]{32}', password):
        return password.lower()

    return hashlib.md5(password.encode('utf-8')).hexdigest()

def validate_nip(nip, row_number=None, required=False):
    prefix = f"Baris {row_number}: " if row_number else ""

    if nip is None or str(nip).strip() == '':
        if required:
            raise ValueError(f"{prefix}NIP wajib diisi")

        return None

    nip = str(nip).strip()

    if not re.fullmatch(r'\d{10,20}', nip):
        raise ValueError(f"{prefix}NIP harus berupa angka 10-20 digit")

    return nip

def validate_nim(nim, row_number=None, required=False):
    prefix = f"Baris {row_number}: " if row_number else ""

    if nim is None or str(nim).strip() == '':
        if required:
            raise ValueError(f"{prefix}NIM wajib diisi")

        return None

    nim = str(nim).strip()

    if len(nim) != 9:
        raise ValueError(f"{prefix}NIM harus 9 karakter")

    return nim

def validate_password(password, row_number=None, required=False):
    prefix = f"Baris {row_number}: " if row_number else ""

    if password is None or str(password).strip() == '':
        if required:
            raise ValueError(f"{prefix}Password wajib diisi")

        return None

    password = str(password)

    if len(password) < 5:
        raise ValueError(f"{prefix}Password minimal 5 karakter")

    return password

# ===============================
# READ EXCEL WITHOUT OPENPYXL FALLBACK
# ===============================
def read_xlsx_rows_with_stdlib(stream):
    stream.seek(0)

    namespace = {
        'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    }

    with zipfile.ZipFile(stream) as archive:
        shared_strings = []

        if 'xl/sharedStrings.xml' in archive.namelist():
            shared_tree = ElementTree.fromstring(
                archive.read('xl/sharedStrings.xml')
            )

            for item in shared_tree.findall('main:si', namespace):
                texts = [
                    node.text or ''
                    for node in item.findall('.//main:t', namespace)
                ]
                shared_strings.append(''.join(texts))

        sheet_tree = ElementTree.fromstring(
            archive.read('xl/worksheets/sheet1.xml')
        )

        parsed_rows = []

        for row in sheet_tree.findall('.//main:row', namespace):
            values = []

            for cell in row.findall('main:c', namespace):
                index = column_index(cell.attrib.get('r', ''))

                while len(values) <= index:
                    values.append(None)

                cell_type = cell.attrib.get('t')
                value_node = cell.find('main:v', namespace)
                inline_node = cell.find('main:is/main:t', namespace)

                if cell_type == 's' and value_node is not None:
                    value = shared_strings[int(value_node.text)]
                elif cell_type == 'inlineStr' and inline_node is not None:
                    value = inline_node.text
                elif value_node is not None:
                    value = value_node.text
                else:
                    value = None

                values[index] = value

            parsed_rows.append(values)

    return parsed_rows


def read_user_excel(file):
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(file.stream, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

    except ImportError:
        rows = read_xlsx_rows_with_stdlib(file.stream)

    if not rows:
        raise ValueError("File Excel kosong")

    header_map = {}

    for index, header in enumerate(rows[0]):
        field = USER_IMPORT_HEADER_ALIASES.get(normalize_import_header(header))

        if field:
            header_map[field] = index

    missing_headers = [
        field
        for field in USER_IMPORT_REQUIRED_FIELDS
        if field not in header_map
    ]

    if missing_headers:
        raise ValueError(
            f"Kolom wajib belum ada: {', '.join(missing_headers)}"
        )

    users = []

    for row_number, row in enumerate(rows[1:], start=2):
        if not row or all(
            value is None or str(value).strip() == ''
            for value in row
        ):
            continue

        item = {}

        for field, index in header_map.items():
            value = row[index] if index < len(row) else None
            item[field] = str(value or '').strip()

        missing_values = [
            field
            for field in USER_IMPORT_REQUIRED_FIELDS
            if not item.get(field)
        ]

        if missing_values:
            raise ValueError(
                f"Baris {row_number}: field wajib kosong "
                f"({', '.join(missing_values)})"
            )

        role = item['role'].strip().lower()

        if role not in ALLOWED_ROLES:
            raise ValueError(
                f"Baris {row_number}: role tidak valid ({role})"
            )

        if role != 'mahasiswa' and not item.get('email'):
            raise ValueError(
                f"Baris {row_number}: email wajib untuk role {role}"
            )

        if role != 'mahasiswa' and not item.get('password'):
            raise ValueError(
                f"Baris {row_number}: password wajib untuk role {role}"
            )

        nip = validate_nip(
            item.get('nip'),
            row_number,
            required=role != 'mahasiswa'
        )
        nim = validate_nim(
            item.get('nim'),
            row_number,
            required=role == 'mahasiswa'
        )
        password = validate_password(
            item.get('password'),
            row_number,
            required=role != 'mahasiswa'
        )

        foto = [
            secure_filename(name.strip())
            for name in item.get('foto', '').split(',')
            if name.strip()
        ]

        users.append({
            'row_number': row_number,
            'nama': item.get('nama'),
            'role': role,
            'nim': nim,
            'nip': nip,
            'prodi': item.get('prodi') or None,
            'kelas': item.get('kelas') or None,
            'email': item.get('email') or None,
            'password': password,
            'foto': foto,
        })

    if not users:
        raise ValueError("Tidak ada data user yang bisa diimport")

    return users


def read_photo_zip(file):
    if not file or not file.filename:
        return {}

    if not file.filename.lower().endswith('.zip'):
        raise ValueError("File foto harus berformat .zip")

    file.stream.seek(0)

    photos = {}

    with zipfile.ZipFile(file.stream) as archive:
        for info in archive.infolist():
            if info.is_dir():
                continue

            filename = secure_filename(os.path.basename(info.filename))

            if not filename:
                continue

            if not allowed_file(filename):
                raise ValueError(
                    f"Format foto di ZIP tidak valid: {filename}"
                )

            photos[filename] = archive.read(info)

    return photos


def save_user_face_bytes(cursor, user_id, face_label, original_filename, content):
    if not allowed_file(original_filename):
        raise ValueError("Format foto harus png, jpg, atau jpeg")

    ext = original_filename.rsplit('.', 1)[1].lower()
    filename = f"{uuid.uuid4().hex}.{ext}"

    user_folder = os.path.join(UPLOAD_FOLDER, face_label)
    os.makedirs(user_folder, exist_ok=True)

    file_path = os.path.join(user_folder, filename)

    with open(file_path, 'wb') as file:
        file.write(content)

    if not os.path.exists(file_path) or os.path.getsize(file_path) == 0:
        raise ValueError(f"Foto gagal disimpan: {original_filename}")

    cursor.execute(
        """
        INSERT INTO user_faces (user_id, image_path, image_name)
        VALUES (%s, %s, %s)
        """,
        (user_id, filename, filename)
    )

    return filename


def save_user_face_files(cursor, user_id, face_label, files):
    if not files:
        return []

    user_folder = os.path.join(UPLOAD_FOLDER, face_label)
    os.makedirs(user_folder, exist_ok=True)

    uploaded_files = []

    for file in files:
        if not allowed_file(file.filename):
            raise ValueError("Format foto harus png, jpg, atau jpeg")

        original_filename = secure_filename(file.filename)
        ext = original_filename.rsplit('.', 1)[1].lower()
        filename = f"{uuid.uuid4().hex}.{ext}"

        file_path = os.path.join(user_folder, filename)
        file.save(file_path)

        if not os.path.exists(file_path) or os.path.getsize(file_path) == 0:
            raise ValueError(f"Foto gagal disimpan: {original_filename}")

        cursor.execute(
            """
            INSERT INTO user_faces (user_id, image_path, image_name)
            VALUES (%s, %s, %s)
            """,
            (user_id, filename, filename)
        )

        uploaded_files.append(filename)

    return uploaded_files


def update_embeddings_and_svm():
    try:
        subprocess.run(['python', 'extract_embedding.py'], check=True)
        subprocess.run(['python', 'training_svm.py'], check=True)

        print("Embeddings and SVM updated successfully.")

    except subprocess.CalledProcessError as e:
        print(f"Error updating embeddings/SVM: {e}")


# ===============================
# GET USERS
# ===============================
@user_bp.route('/users', methods=['GET'])
def get_users():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT id, kode, nama, face_label, role,
               nim, nip, prodi, kelas,
               email
        FROM users
        ORDER BY id DESC
    """)

    users = cursor.fetchall()

    for user in users:
        cursor.execute(
            """
            SELECT image_name
            FROM user_faces
            WHERE user_id = %s
            """,
            (user['id'],)
        )

        faces = cursor.fetchall()
        user['user_faces'] = faces

    cursor.close()
    conn.close()

    return jsonify(users), 200


# ===============================
# GENERATE KODE
# ===============================
def generate_user_code(role, cursor):
    prefix_map = {
        'kepala_lab': 'KL',
        'teknisi': 'TK',
        'dosen': 'DS',
        'sarpras': 'SP',
        'mahasiswa': 'MH'
    }

    prefix = prefix_map.get(role)

    if not prefix:
        raise ValueError("Role tidak valid untuk generate kode")

    cursor.execute(
        """
        SELECT kode
        FROM users
        WHERE role = %s
        ORDER BY id DESC
        LIMIT 1
        """,
        (role,)
    )

    last = cursor.fetchone()

    if last and last[0]:
        last_num = int(last[0][2:])
        new_num = last_num + 1
    else:
        new_num = 1

    return f"{prefix}{str(new_num).zfill(4)}"


@user_bp.route('/users/generate_kode', methods=['GET'])
def api_generate_kode():
    role = request.args.get('role')

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        kode = generate_user_code(role, cursor)

    except Exception as e:
        return jsonify({"message": str(e)}), 400

    finally:
        cursor.close()
        conn.close()

    return jsonify({"kode": kode}), 200


# ===============================
# CREATE USER
# ===============================
@user_bp.route('/users', methods=['POST'])
def create_user():
    files = []

    if request.content_type and request.content_type.startswith('multipart/form-data'):
        nama = request.form.get('nama')
        role = request.form.get('role')
        nim = request.form.get('nim')
        nip = request.form.get('nip')
        prodi = request.form.get('prodi')
        kelas = request.form.get('kelas')
        email = request.form.get('email')
        password = request.form.get('password')
        files = get_uploaded_files()

    else:
        data = request.get_json() or {}
        nama = data.get('nama')
        role = data.get('role')
        nim = data.get('nim')
        nip = data.get('nip')
        prodi = data.get('prodi')
        kelas = data.get('kelas')
        email = data.get('email')
        password = data.get('password')

    role = str(role or '').strip().lower()

    if not nama:
        return jsonify({"message": "Nama wajib diisi"}), 400

    if role not in ALLOWED_ROLES:
        return jsonify({"message": "Role tidak valid"}), 400

    try:
        nim = validate_nim(nim, required=role == 'mahasiswa')
        nip = validate_nip(nip, required=role != 'mahasiswa')
        password = validate_password(password, required=role != 'mahasiswa')
    except ValueError as e:
        return jsonify({"message": str(e)}), 400

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        if role != 'mahasiswa':
            if not email:
                return jsonify({"message": "Email wajib diisi"}), 400

            cursor.execute(
                "SELECT id FROM users WHERE email = %s",
                (email,)
            )

            if cursor.fetchone():
                return jsonify({
                    "message": f"Email '{email}' sudah terdaftar"
                }), 400

            password = hash_password_md5(password)

        if role == 'mahasiswa':
            email = None
            password = None
            nip = None

        kode = generate_user_code(role, cursor)
        face_label = generate_face_label(nama)

        cursor.execute(
            """
            INSERT INTO users
            (kode, nama, face_label, role, nim, nip, prodi, kelas,
             email, password)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                kode, nama, face_label, role,
                nim, nip, prodi, kelas,
                email, password
            )
        )

        user_id = cursor.lastrowid

        uploaded_files = save_user_face_files(
            cursor,
            user_id,
            face_label,
            files
        )

        conn.commit()

        if uploaded_files:
            update_embeddings_and_svm()

    except Exception as e:
        conn.rollback()
        return jsonify({
            "message": f"Gagal membuat user: {str(e)}"
        }), 500

    finally:
        cursor.close()
        conn.close()

    return jsonify({
        "message": "User berhasil dibuat",
        "user_id": user_id,
        "kode": kode,
        "face_label": face_label,
        "files_uploaded": uploaded_files
    }), 201

# ===============================
# IMPORT USER DARI EXCEL + ZIP FOTO
# ===============================
@user_bp.route('/users/import', methods=['POST'])
def import_users():
    excel_file = request.files.get('excel') or request.files.get('file')
    photos_zip = request.files.get('photos_zip')

    if not excel_file or not excel_file.filename:
        return jsonify({"message": "File Excel wajib diupload"}), 400

    if not excel_file.filename.lower().endswith('.xlsx'):
        return jsonify({"message": "Format file harus .xlsx"}), 400

    try:
        rows = read_user_excel(excel_file)
        photos = read_photo_zip(photos_zip)

    except Exception as e:
        return jsonify({"message": str(e)}), 400

    conn = get_db_connection()
    cursor = conn.cursor()

    created_users = []

    try:
        for row in rows:
            nama = row['nama']
            role = row['role']
            nim = row['nim']
            nip = row['nip']
            prodi = row['prodi']
            kelas = row['kelas']
            email = row['email']
            password = row['password']

            if role == 'mahasiswa':
                email = None
                password = None
            else:
                password = hash_password_md5(password)

            kode = generate_user_code(role, cursor)
            face_label = generate_face_label(nama)

            cursor.execute(
                """
                INSERT INTO users
                (kode, nama, face_label, role, nim, nip, prodi, kelas,
                 email, password)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    kode, nama, face_label, role,
                    nim, nip, prodi, kelas,
                    email, password
                )
            )

            user_id = cursor.lastrowid
            uploaded_files = []

            for photo_name in row.get('foto', []):
                if photo_name not in photos:
                    raise ValueError(
                        f"Baris {row['row_number']}: foto '{photo_name}' "
                        f"tidak ditemukan di ZIP"
                    )

                saved_name = save_user_face_bytes(
                    cursor,
                    user_id,
                    face_label,
                    photo_name,
                    photos[photo_name]
                )

                uploaded_files.append(saved_name)

            created_users.append({
                "user_id": user_id,
                "kode": kode,
                "nama": nama,
                "face_label": face_label,
                "files_uploaded": uploaded_files
            })

        conn.commit()

        update_embeddings_and_svm()

    except Exception as e:
        conn.rollback()
        return jsonify({
            "message": f"Gagal import user: {str(e)}"
        }), 500

    finally:
        cursor.close()
        conn.close()

    return jsonify({
        "message": f"Berhasil import {len(created_users)} user",
        "imported": len(created_users),
        "users": created_users
    }), 201

# ===============================
# UPLOAD FOTO USER
# ===============================
@user_bp.route('/users/<int:user_id>/upload_faces', methods=['POST'])
def upload_user_faces(user_id):
    files = get_uploaded_files()

    if not files:
        return jsonify({"message": "No selected files"}), 400

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute(
            "SELECT face_label FROM users WHERE id = %s",
            (user_id,)
        )

        user = cursor.fetchone()

        if not user:
            return jsonify({"message": "User not found"}), 404

        face_label = user['face_label']

        uploaded_files = save_user_face_files(
            cursor,
            user_id,
            face_label,
            files
        )

        conn.commit()

        if uploaded_files:
            update_embeddings_and_svm()

    except Exception as e:
        conn.rollback()
        return jsonify({
            "message": f"Gagal upload foto: {str(e)}"
        }), 500

    finally:
        cursor.close()
        conn.close()

    return jsonify({
        "message": "Files uploaded successfully",
        "files": uploaded_files
    }), 201


# ===============================
# PREVIEW FOTO
# ===============================
@user_bp.route('/uploads/<face_label>/<filename>', methods=['GET'])
def uploaded_file(face_label, filename):
    folder = os.path.join(UPLOAD_FOLDER, face_label)
    file_path = os.path.join(folder, filename)

    print("Looking for file:", file_path)

    if not os.path.exists(file_path):
        return jsonify({
            "message": f"File not found: {file_path}"
        }), 404

    return send_from_directory(folder, filename)

# ===============================
# UPDATE USER
# ===============================
@user_bp.route('/users/<int:user_id>', methods=['PUT'])
def update_user(user_id):
    data = request.get_json() or {}

    nama = data.get('nama')
    role = data.get('role')
    nim = data.get('nim')
    nip = data.get('nip')
    prodi = data.get('prodi')
    kelas = data.get('kelas')
    email = data.get('email')
    password = data.get('password')

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute(
            "SELECT * FROM users WHERE id = %s",
            (user_id,)
        )

        user = cursor.fetchone()

        if not user:
            return jsonify({"message": "User tidak ditemukan"}), 404

        nama = nama if nama is not None else user.get('nama')
        role = role if role is not None else user.get('role')
        nim = nim if nim is not None else user.get('nim')
        nip = nip if nip is not None else user.get('nip')
        prodi = prodi if prodi is not None else user.get('prodi')
        kelas = kelas if kelas is not None else user.get('kelas')
        email = email if email is not None else user.get('email')

        role = str(role or '').strip().lower()

        if role not in ALLOWED_ROLES:
            return jsonify({"message": "Role tidak valid"}), 400

        try:
            nim = validate_nim(nim, required=role == 'mahasiswa')
            nip = validate_nip(nip, required=role != 'mahasiswa')
            password = validate_password(password, required=False)
        except ValueError as e:
            return jsonify({"message": str(e)}), 400

        if role == 'mahasiswa':
            email = None
            password = None
            nip = None
        else:
            if password is not None and str(password).strip() != '':
                password = hash_password_md5(password)
            else:
                password = user.get('password')

        cursor.execute(
            """
            UPDATE users SET
                nama = %s,
                role = %s,
                nim = %s,
                nip = %s,
                prodi = %s,
                kelas = %s,
                email = %s,
                password = %s
            WHERE id = %s
            """,
            (
                nama, role, nim, nip,
                prodi, kelas, email, password,
                user_id
            )
        )

        conn.commit()

        cursor.execute(
            """
            SELECT id, kode, nama, face_label, role,
                   nim, nip, prodi, kelas,
                   email
            FROM users
            WHERE id = %s
            """,
            (user_id,)
        )

        updated_user = cursor.fetchone()

    except Exception as e:
        conn.rollback()
        return jsonify({
            "message": f"Gagal update user: {str(e)}"
        }), 500

    finally:
        cursor.close()
        conn.close()

    return jsonify(updated_user), 200

# ===============================
# DELETE USER
# ===============================
@user_bp.route('/users/<int:user_id>', methods=['DELETE'])
def delete_user(user_id):
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute(
            "SELECT face_label FROM users WHERE id = %s",
            (user_id,)
        )

        user = cursor.fetchone()

        if not user:
            return jsonify({"message": "User tidak ditemukan"}), 404

        face_label = user[0]

        cursor.execute(
            "DELETE FROM user_faces WHERE user_id = %s",
            (user_id,)
        )

        folder = os.path.join(UPLOAD_FOLDER, face_label)

        if os.path.exists(folder) and os.path.isdir(folder):
            shutil.rmtree(folder, ignore_errors=True)

        cursor.execute(
            "DELETE FROM users WHERE id = %s",
            (user_id,)
        )

        conn.commit()

    except Exception as e:
        conn.rollback()
        return jsonify({
            "message": f"Gagal hapus user: {str(e)}"
        }), 500

    finally:
        cursor.close()
        conn.close()

    return jsonify({"message": "User berhasil dihapus"}), 200

# ===============================
# GET DOSEN OPTIONS
# ===============================
@user_bp.route('/users/dosen-options', methods=['GET'])
def get_dosen_options():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            SELECT id, kode, nama, nip, email
            FROM users
            WHERE role = 'dosen'
            ORDER BY nama ASC
        """)

        dosen = cursor.fetchall()

    except Exception as e:
        return jsonify({
            "message": f"Gagal mengambil data dosen: {str(e)}"
        }), 500

    finally:
        cursor.close()
        conn.close()

    return jsonify(dosen), 200
