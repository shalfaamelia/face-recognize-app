from flask import Blueprint, request, jsonify
from db import get_db_connection
import re
import zipfile
from xml.etree import ElementTree

jadwal_bp = Blueprint('jadwal', __name__)

JADWAL_REQUIRED_FIELDS = [
    'nama',
    'kelas',
    'hari',
    'jam_mulai',
    'jam_selesai'
]

JADWAL_IMPORT_REQUIRED_FIELDS = [
    'nama',
    'dosen',
    'kelas',
    'hari',
    'jam_mulai',
    'jam_selesai'
]

JADWAL_IMPORT_HEADER_ALIASES = {
    'kode': 'kode',
    'kode jadwal': 'kode',

    'nama': 'nama',
    'mata kuliah': 'nama',
    'nama mata kuliah': 'nama',
    'praktikum': 'nama',
    'nama praktikum': 'nama',

    'dosen': 'dosen',
    'nama dosen': 'dosen',

    'nip': 'nip',
    'nip dosen': 'nip',

    'kelas': 'kelas',

    'hari': 'hari',

    'jam mulai': 'jam_mulai',
    'jam_mulai': 'jam_mulai',
    'mulai': 'jam_mulai',

    'jam selesai': 'jam_selesai',
    'jam_selesai': 'jam_selesai',
    'selesai': 'jam_selesai',
}

HARI_VALID = {
    'senin',
    'selasa',
    'rabu',
    'kamis',
    'jumat',
    'sabtu',
    'minggu'
}


# ===============================
# HELPER
# ===============================
def normalize_header(value):
    return str(value or '').strip().lower().replace('\n', ' ')


def normalize_text(value):
    return str(value or '').strip()


def normalize_hari(value):
    value = normalize_text(value).lower()

    mapping = {
        'senin': 'Senin',
        'selasa': 'Selasa',
        'rabu': 'Rabu',
        'kamis': 'Kamis',
        'jumat': 'Jumat',
        "jum'at": 'Jumat',
        'jum’at': 'Jumat',
        'sabtu': 'Sabtu',
        'minggu': 'Minggu',
    }

    return mapping.get(value, normalize_text(value))


def normalize_time(value):
    if value is None:
        return None

    value = str(value).strip()

    if not value:
        return None

    # Format dari Excel kadang berupa angka pecahan hari: 0.3333 = 08:00
    try:
        number_value = float(value)
        if 0 <= number_value < 1:
            total_minutes = round(number_value * 24 * 60)
            hour = total_minutes // 60
            minute = total_minutes % 60
            return f"{hour:02d}:{minute:02d}:00"
    except ValueError:
        pass

    # Format HH:MM
    match = re.match(r'^(\d{1,2}):(\d{2})(?::(\d{2}))?$', value)
    if match:
        hour = int(match.group(1))
        minute = int(match.group(2))
        second = int(match.group(3) or 0)

        if hour > 23 or minute > 59 or second > 59:
            raise ValueError(f"Format jam tidak valid: {value}")

        return f"{hour:02d}:{minute:02d}:{second:02d}"

    raise ValueError(f"Format jam tidak valid: {value}. Gunakan format HH:MM")


def format_time_for_response(value):
    if value is None:
        return None

    value = str(value)

    if len(value) >= 5:
        return value[:5]

    return value


def column_index(cell_reference):
    letters = ''.join(re.findall(r'[A-Z]+', cell_reference.upper()))
    index = 0

    for char in letters:
        index = index * 26 + (ord(char) - ord('A') + 1)

    return index - 1


def read_xlsx_rows_with_stdlib(stream):
    stream.seek(0)

    namespace = {
        'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    }

    with zipfile.ZipFile(stream) as archive:
        shared_strings = []

        if 'xl/sharedStrings.xml' in archive.namelist():
            shared_tree = ElementTree.fromstring(
                archive.read('xl/sharedStrings.xml')
            )

            for item in shared_tree.findall('main:si', namespace):
                texts = [
                    node.text or ''
                    for node in item.findall('.//main:t', namespace)
                ]
                shared_strings.append(''.join(texts))

        sheet_tree = ElementTree.fromstring(
            archive.read('xl/worksheets/sheet1.xml')
        )

        parsed_rows = []

        for row in sheet_tree.findall('.//main:row', namespace):
            values = []

            for cell in row.findall('main:c', namespace):
                index = column_index(cell.attrib.get('r', ''))

                while len(values) <= index:
                    values.append(None)

                cell_type = cell.attrib.get('t')
                value_node = cell.find('main:v', namespace)
                inline_node = cell.find('main:is/main:t', namespace)

                if cell_type == 's' and value_node is not None:
                    value = shared_strings[int(value_node.text)]
                elif cell_type == 'inlineStr' and inline_node is not None:
                    value = inline_node.text
                elif value_node is not None:
                    value = value_node.text
                else:
                    value = None

                values[index] = value

            parsed_rows.append(values)

    return parsed_rows


def read_jadwal_excel(file):
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(file.stream, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

    except ImportError:
        rows = read_xlsx_rows_with_stdlib(file.stream)

    if not rows:
        raise ValueError("File Excel kosong")

    header_map = {}

    for index, header in enumerate(rows[0]):
        field = JADWAL_IMPORT_HEADER_ALIASES.get(normalize_header(header))

        if field:
            header_map[field] = index

    missing_headers = [
        field
        for field in JADWAL_IMPORT_REQUIRED_FIELDS
        if field not in header_map
    ]

    if missing_headers:
        raise ValueError(
            f"Kolom wajib belum ada: {', '.join(missing_headers)}"
        )

    schedules = []

    for row_number, row in enumerate(rows[1:], start=2):
        if not row or all(
            value is None or str(value).strip() == ''
            for value in row
        ):
            continue

        item = {}

        for field, index in header_map.items():
            value = row[index] if index < len(row) else None
            item[field] = normalize_text(value)

        missing_values = [
            field
            for field in JADWAL_IMPORT_REQUIRED_FIELDS
            if not item.get(field)
        ]

        if missing_values:
            raise ValueError(
                f"Baris {row_number}: field wajib kosong "
                f"({', '.join(missing_values)})"
            )

        hari = normalize_hari(item.get('hari'))

        if hari.lower() not in HARI_VALID:
            raise ValueError(
                f"Baris {row_number}: hari tidak valid ({item.get('hari')})"
            )

        jam_mulai = normalize_time(item.get('jam_mulai'))
        jam_selesai = normalize_time(item.get('jam_selesai'))

        schedules.append({
            'row_number': row_number,
            'kode': item.get('kode') or None,
            'nama': item.get('nama'),
            'dosen': item.get('dosen'),
            'nip': item.get('nip') or None,
            'kelas': item.get('kelas'),
            'hari': hari,
            'jam_mulai': jam_mulai,
            'jam_selesai': jam_selesai,
        })

    if not schedules:
        raise ValueError("Tidak ada data jadwal yang bisa diimport")

    return schedules


def find_dosen(cursor, dosen_user_id=None, dosen=None, nip=None):
    if dosen_user_id:
        cursor.execute("""
            SELECT id, nama, nip
            FROM users
            WHERE id = %s AND role = 'dosen'
        """, (dosen_user_id,))

        result = cursor.fetchone()

        if result:
            return result

    if nip:
        cursor.execute("""
            SELECT id, nama, nip
            FROM users
            WHERE nip = %s AND role = 'dosen'
            LIMIT 1
        """, (nip,))

        result = cursor.fetchone()

        if result:
            return result

    if dosen:
        cursor.execute("""
            SELECT id, nama, nip
            FROM users
            WHERE LOWER(nama) = LOWER(%s) AND role = 'dosen'
            LIMIT 1
        """, (dosen,))

        result = cursor.fetchone()

        if result:
            return result

    return None


def validate_jadwal_payload(data):
    missing = [
        field
        for field in JADWAL_REQUIRED_FIELDS
        if not normalize_text(data.get(field))
    ]

    if missing:
        raise ValueError(f"Field wajib belum lengkap: {', '.join(missing)}")

    hari = normalize_hari(data.get('hari'))

    if hari.lower() not in HARI_VALID:
        raise ValueError("Hari tidak valid")

    jam_mulai = normalize_time(data.get('jam_mulai'))
    jam_selesai = normalize_time(data.get('jam_selesai'))

    return {
        'kode': normalize_text(data.get('kode')) or None,
        'nama': normalize_text(data.get('nama')),
        'dosen_user_id': data.get('dosen_user_id') or None,
        'dosen': normalize_text(data.get('dosen')) or None,
        'nip': normalize_text(data.get('nip')) or None,
        'kelas': normalize_text(data.get('kelas')),
        'hari': hari,
        'jam_mulai': jam_mulai,
        'jam_selesai': jam_selesai,
    }


def format_jadwal_rows(rows):
    for row in rows:
        row['jam_mulai'] = format_time_for_response(row.get('jam_mulai'))
        row['jam_selesai'] = format_time_for_response(row.get('jam_selesai'))

    return rows


# ===============================
# GET JADWAL
# ===============================
@jadwal_bp.route('/', methods=['GET'])
def get_jadwal():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            SELECT 
                jp.id,
                jp.kode,
                jp.nama,
                jp.dosen_user_id,
                jp.dosen,
                jp.nip,
                jp.kelas,
                jp.hari,
                jp.jam_mulai,
                jp.jam_selesai,
                u.nama AS dosen_nama_user,
                u.nip AS dosen_nip_user
            FROM jadwal_praktikum jp
            LEFT JOIN users u ON jp.dosen_user_id = u.id
            ORDER BY 
                FIELD(jp.hari, 'Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu'),
                jp.jam_mulai ASC
        """)

        rows = cursor.fetchall()
        rows = format_jadwal_rows(rows)

    except Exception as e:
        return jsonify({
            "message": f"Gagal mengambil jadwal: {str(e)}"
        }), 500

    finally:
        cursor.close()
        conn.close()

    return jsonify(rows), 200


# ===============================
# CREATE JADWAL
# ===============================
@jadwal_bp.route('/', methods=['POST'])
def create_jadwal():
    data = request.get_json() or {}

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        payload = validate_jadwal_payload(data)

        dosen = find_dosen(
            cursor,
            payload.get('dosen_user_id'),
            payload.get('dosen'),
            payload.get('nip')
        )

        if not dosen:
            raise ValueError("Dosen tidak ditemukan. Pastikan dosen sudah ada di menu User dengan role dosen.")

        cursor.execute("""
            INSERT INTO jadwal_praktikum
            (
                kode,
                nama,
                dosen_user_id,
                dosen,
                nip,
                kelas,
                hari,
                jam_mulai,
                jam_selesai
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            payload['kode'],
            payload['nama'],
            dosen['id'],
            dosen['nama'],
            dosen['nip'],
            payload['kelas'],
            payload['hari'],
            payload['jam_mulai'],
            payload['jam_selesai'],
        ))

        conn.commit()

        jadwal_id = cursor.lastrowid

    except Exception as e:
        conn.rollback()
        return jsonify({
            "message": f"Gagal membuat jadwal: {str(e)}"
        }), 500

    finally:
        cursor.close()
        conn.close()

    return jsonify({
        "message": "Jadwal berhasil dibuat",
        "id": jadwal_id
    }), 201


# ===============================
# UPDATE JADWAL
# ===============================
@jadwal_bp.route('/<int:jadwal_id>', methods=['PUT'])
def update_jadwal(jadwal_id):
    data = request.get_json() or {}

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            SELECT *
            FROM jadwal_praktikum
            WHERE id = %s
        """, (jadwal_id,))

        existing = cursor.fetchone()

        if not existing:
            return jsonify({"message": "Jadwal tidak ditemukan"}), 404

        merged = {
            'kode': data.get('kode', existing.get('kode')),
            'nama': data.get('nama', existing.get('nama')),
            'dosen_user_id': data.get('dosen_user_id', existing.get('dosen_user_id')),
            'dosen': data.get('dosen', existing.get('dosen')),
            'nip': data.get('nip', existing.get('nip')),
            'kelas': data.get('kelas', existing.get('kelas')),
            'hari': data.get('hari', existing.get('hari')),
            'jam_mulai': data.get('jam_mulai', existing.get('jam_mulai')),
            'jam_selesai': data.get('jam_selesai', existing.get('jam_selesai')),
        }

        payload = validate_jadwal_payload(merged)

        dosen = find_dosen(
            cursor,
            payload.get('dosen_user_id'),
            payload.get('dosen'),
            payload.get('nip')
        )

        if not dosen:
            raise ValueError("Dosen tidak ditemukan. Pastikan dosen sudah ada di menu User dengan role dosen.")

        cursor.execute("""
            UPDATE jadwal_praktikum SET
                kode = %s,
                nama = %s,
                dosen_user_id = %s,
                dosen = %s,
                nip = %s,
                kelas = %s,
                hari = %s,
                jam_mulai = %s,
                jam_selesai = %s
            WHERE id = %s
        """, (
            payload['kode'],
            payload['nama'],
            dosen['id'],
            dosen['nama'],
            dosen['nip'],
            payload['kelas'],
            payload['hari'],
            payload['jam_mulai'],
            payload['jam_selesai'],
            jadwal_id
        ))

        conn.commit()

    except Exception as e:
        conn.rollback()
        return jsonify({
            "message": f"Gagal update jadwal: {str(e)}"
        }), 500

    finally:
        cursor.close()
        conn.close()

    return jsonify({"message": "Jadwal berhasil diupdate"}), 200


# ===============================
# DELETE JADWAL
# ===============================
@jadwal_bp.route('/<int:jadwal_id>', methods=['DELETE'])
def delete_jadwal(jadwal_id):
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            DELETE FROM jadwal_praktikum
            WHERE id = %s
        """, (jadwal_id,))

        conn.commit()

        if cursor.rowcount == 0:
            return jsonify({"message": "Jadwal tidak ditemukan"}), 404

    except Exception as e:
        conn.rollback()
        return jsonify({
            "message": f"Gagal hapus jadwal: {str(e)}"
        }), 500

    finally:
        cursor.close()
        conn.close()

    return jsonify({"message": "Jadwal berhasil dihapus"}), 200


# ===============================
# IMPORT JADWAL EXCEL
# ===============================
@jadwal_bp.route('/import', methods=['POST'])
def import_jadwal():
    excel_file = request.files.get('file') or request.files.get('excel')

    if not excel_file or not excel_file.filename:
        return jsonify({"message": "File Excel wajib diupload"}), 400

    if not excel_file.filename.lower().endswith('.xlsx'):
        return jsonify({"message": "Format file harus .xlsx"}), 400

    try:
        rows = read_jadwal_excel(excel_file)

    except Exception as e:
        return jsonify({"message": str(e)}), 400

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    imported = []

    try:
        for row in rows:
            dosen = find_dosen(
                cursor,
                dosen_user_id=None,
                dosen=row.get('dosen'),
                nip=row.get('nip')
            )

            if not dosen:
                raise ValueError(
                    f"Baris {row['row_number']}: dosen '{row.get('dosen')}' "
                    f"dengan NIP '{row.get('nip') or '-'}' tidak ditemukan di tabel users"
                )

            cursor.execute("""
                INSERT INTO jadwal_praktikum
                (
                    kode,
                    nama,
                    dosen_user_id,
                    dosen,
                    nip,
                    kelas,
                    hari,
                    jam_mulai,
                    jam_selesai
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                row['kode'],
                row['nama'],
                dosen['id'],
                dosen['nama'],
                dosen['nip'],
                row['kelas'],
                row['hari'],
                row['jam_mulai'],
                row['jam_selesai'],
            ))

            imported.append({
                "row_number": row['row_number'],
                "nama": row['nama'],
                "dosen": dosen['nama'],
                "nip": dosen['nip'],
                "kelas": row['kelas'],
                "hari": row['hari'],
            })

        conn.commit()

    except Exception as e:
        conn.rollback()
        return jsonify({
            "message": f"Gagal import jadwal: {str(e)}"
        }), 500

    finally:
        cursor.close()
        conn.close()

    return jsonify({
        "message": f"Berhasil import {len(imported)} jadwal",
        "imported": len(imported),
        "data": imported
    }), 201